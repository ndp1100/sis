#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 30 13:06:25 2020

@author: PhuongND
"""
from glob import glob
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib
import urllib.request
import validators
import os
import pathlib
import pandas as pd

path = pathlib.Path().absolute()
print(path)
path = os.path.join(path, 'static\img')
print(path)


def is_not_blank(s):
    return bool(s and s.strip())


def downloadimg(msp, linkImage):
    if is_not_blank(linkImage):
        if validators.url(linkImage):
            filedownloaded = os.path.join(path, msp + '.jpg')
            urllib.request.urlretrieve(linkImage, filedownloaded)
    return


df = pd.read_excel('AllProduct.xlsx')


def GetGiaBanFromMSP(msp):
    # print('Search Gia MSP ', msp)
    df_filterMSP = df.loc[df['Mã sản phẩm'] == msp, ['Giá bán', 'Mã danh mục']]
    currency = 0
    maDanhMuc = str('None')
    if df_filterMSP.values.size != 0:
        currency = "{:,.0f} VND".format(df_filterMSP.iloc[0]['Giá bán'])
        maDanhMuc = df_filterMSP.iloc[0]['Mã danh mục']
    # print(currency)

    return currency, maDanhMuc

#
# tien, ma = GetGiaBanFromMSP('TTDB-BJ-0167')
# print(tien)
# print(ma)

excelFiles = glob('*.xlsx', recursive=False)


def GetInforFromMSP(msp):
    for j in excelFiles:
        # print(j)
        wb = load_workbook(filename=j)
        ws = wb.active
        # print(ws.title) 
        # print(ws['F2'].value)
        try:
            # for value in ws.iter_rows(2, ws.max_row, 1,ws.max_column, values_only=True):
            #     # print(value[0], 'row', i) #ma san pham
            #     return value[0]

            for row in ws.iter_rows("A"):
                for cell in row:
                    if cell.value == msp:
                        return ws.cell(row=cell.row, column=4).value
        except Exception as e:
            pass


def DownloadImgFromExcelFile():
    for j in excelFiles:
        # print(j)
        wb = load_workbook(filename=j)
        ws = wb.active
        print(ws.title)
        try:
            i = 1
            for value in ws.iter_rows(2, ws.max_row, 1, 2, values_only=True):
                print(value[0], 'row', i)  # ma san pham
                print(value[1])  # link image
                downloadimg(value[0], value[1])
                i = i + 1
        except Exception as e:
            print("Exception: ", value[0], e)
            pass
